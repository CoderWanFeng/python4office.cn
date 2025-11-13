#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel公式批量应用工具
功能：批量应用公式到Excel文件的指定单元格，支持常见数学、统计和文本公式
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from PySide6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QComboBox, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox, QSpinBox, QMessageBox
from PySide6.QtCore import QTimer, Qt

# 添加父目录到路径，以便导入common模块
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)

from common.ui_base import BaseMainWindow, FileSelectionWidget, OperationWidget


class ExcelFormulaWindow(BaseMainWindow):
    """Excel公式批量应用工具主窗口"""
    
    def __init__(self):
        super().__init__("Excel公式批量应用工具", 1000, 700)
        self.workbook = None  # 存储当前工作簿
        self.worksheet = None  # 存储当前工作表
        self.worksheets = []  # 存储所有工作表名
        
    def setup_content(self):
        """设置内容区域"""
        # 文件选择组件
        self.file_selector = FileSelectionWidget("选择要应用公式的Excel文件", multi_selection=False)
        self.file_selector.select_files = self.load_excel_file  # 重写选择文件方法
        self.content_layout.addWidget(self.file_selector)
        
        # 工作表和数据预览
        self.preview_widget = PreviewWidget()
        self.preview_widget.worksheet_changed.connect(self.change_worksheet)
        self.content_layout.addWidget(self.preview_widget)
        
        # 公式应用组件
        self.formula_widget = FormulaApplicationWidget()
        self.content_layout.addWidget(self.formula_widget)
        
        # 操作区域
        self.operation = OperationWidget("公式应用操作")
        self.operation.start_operation = self.apply_formulas
        self.operation.signals.progress.connect(self.update_progress)
        self.operation.signals.finished.connect(self.operation_finished)
        self.content_layout.addWidget(self.operation)
        
    def load_excel_file(self):
        """加载Excel文件"""
        files = QFileDialog.getOpenFileNames(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)")[0]
        
        if not files:
            return
            
        file_path = files[0]
        self.file_selector.file_paths = files
        self.file_selector.update_display()
        
        try:
            # 加载工作簿
            self.workbook = load_workbook(file_path)
            self.worksheets = self.workbook.sheetnames
            
            # 选择第一个工作表
            self.worksheet = self.workbook[self.worksheets[0]]
            
            # 更新预览
            self.preview_widget.update_workbook(self.workbook, self.worksheets)
            
            # 更新状态
            self.statusBar().showMessage(f"已加载文件: {os.path.basename(file_path)}, 共 {len(self.worksheets)} 个工作表")
            
        except Exception as e:
            self.show_error(f"加载Excel文件失败: {str(e)}")
            
    def change_worksheet(self, sheet_name):
        """切换工作表"""
        if sheet_name in self.worksheets:
            self.worksheet = self.workbook[sheet_name]
            self.preview_widget.update_worksheet(self.worksheet)
            
    def apply_formulas(self):
        """应用公式"""
        if self.workbook is None or self.worksheet is None:
            self.show_warning("请先加载Excel文件")
            return
            
        try:
            # 获取公式应用配置
            formula_configs = self.formula_widget.get_formula_configs()
            
            if not formula_configs:
                self.show_warning("请至少添加一个公式配置")
                return
                
            # 获取保存路径
            save_path = self.get_save_path("保存应用公式后的Excel文件", "Excel文件 (*.xlsx)", "公式应用结果.xlsx")
            if not save_path:
                return
                
            # 开始处理
            self.operation.set_running_state(True)
            self.operation.signals.progress.emit(0, 100, "开始应用公式...")
            
            # 使用定时器异步执行
            QTimer.singleShot(100, lambda: self.process_formulas(formula_configs, save_path))
            
        except Exception as e:
            self.show_error(f"应用公式过程中出错: {str(e)}")
            
    def process_formulas(self, formula_configs, save_path):
        """处理公式应用"""
        try:
            total_configs = len(formula_configs)
            
            for i, config in enumerate(formula_configs):
                if not self.operation.is_running:
                    break
                    
                self.operation.signals.progress.emit(
                    int((i + 1) / total_configs * 80), 
                    100, 
                    f"正在应用公式 {i+1}/{total_configs}"
                )
                
                # 提取配置参数
                column_range = config["column_range"]
                formula_type = config["formula_type"]
                custom_formula = config.get("custom_formula", "")
                apply_mode = config["apply_mode"]
                start_row = config["start_row"]
                end_row = config.get("end_row", None)
                
                # 生成公式
                formula = self.generate_formula(
                    formula_type, custom_formula, column_range, apply_mode, start_row, end_row
                )
                
                # 应用公式
                self.apply_formula_to_cells(formula, apply_mode, start_row, end_row)
            
            # 保存工作簿
            self.operation.signals.progress.emit(80, 100, "正在保存Excel文件...")
            self.workbook.save(save_path)
            
            # 完成
            if self.operation.is_running:
                self.operation.signals.finished.emit(True, f"公式应用完成，已保存到: {save_path}")
            else:
                self.operation.signals.finished.emit(False, "用户取消操作")
                
        except Exception as e:
            self.operation.signals.finished.emit(False, f"应用公式过程中出错: {str(e)}")
            
    def generate_formula(self, formula_type, custom_formula, column_range, apply_mode, start_row, end_row):
        """生成Excel公式"""
        if formula_type == "自定义":
            return custom_formula
            
        # 解析列范围
        if ":" in column_range:
            start_col, end_col = column_range.split(":")
        else:
            start_col = end_col = column_range
            
        # 根据公式类型生成公式
        if formula_type == "求和":
            if apply_mode == "整列":
                return f"=SUM({column_range})"
            else:
                return f"=SUM({start_col}{start_row}:{end_col}{end_row})"
                
        elif formula_type == "平均值":
            if apply_mode == "整列":
                return f"=AVERAGE({column_range})"
            else:
                return f"=AVERAGE({start_col}{start_row}:{end_col}{end_row})"
                
        elif formula_type == "计数":
            if apply_mode == "整列":
                return f"=COUNT({column_range})"
            else:
                return f"=COUNT({start_col}{start_row}:{end_col}{end_row})"
                
        elif formula_type == "最大值":
            if apply_mode == "整列":
                return f"=MAX({column_range})"
            else:
                return f"=MAX({start_col}{start_row}:{end_col}{end_row})"
                
        elif formula_type == "最小值":
            if apply_mode == "整列":
                return f"=MIN({column_range})"
            else:
                return f"=MIN({start_col}{start_row}:{end_col}{end_row})"
                
        elif formula_type == "乘积":
            if apply_mode == "整列":
                return f"=PRODUCT({column_range})"
            else:
                return f"=PRODUCT({start_col}{start_row}:{end_col}{end_row})"
                
        elif formula_type == "条件计数":
            # 生成条件计数公式，需要额外的条件参数
            if apply_mode == "整列":
                return f"=COUNTIF({column_range},\">0\")"  # 默认条件大于0
            else:
                return f"=COUNTIF({start_col}{start_row}:{end_col}{end_row},\">0\")"
                
        elif formula_type == "文本连接":
            # 生成文本连接公式，连接两列
            if ":" in column_range:
                return f"=CONCATENATE({start_col}{start_row},{end_col}{start_row})"
            else:
                return f"=CONCATENATE({column_range}{start_row},{column_range}{end_row})"
                
        else:
            return custom_formula
            
    def apply_formula_to_cells(self, formula, apply_mode, start_row, end_row):
        """应用公式到单元格"""
        if apply_mode == "单个单元格":
            # 应用到单个单元格
            cell = self.worksheet.cell(row=start_row, column=1)  # 默认第一列
            cell.value = formula
        elif apply_mode == "整列":
            # 应用到整列
            for row in range(1, self.worksheet.max_row + 1):
                cell = self.worksheet.cell(row=row, column=1)  # 默认第一列
                cell.value = formula
        elif apply_mode == "指定范围":
            # 应用到指定范围
            target_col = 1  # 默认第一列
            for row in range(start_row, end_row + 1):
                if row <= self.worksheet.max_row:
                    cell = self.worksheet.cell(row=row, column=target_col)
                    cell.value = formula
                    
    def update_progress(self, current, total, status):
        """更新进度"""
        self.operation.progress.update_progress(current, total, status)
        
    def operation_finished(self, success, message):
        """操作完成"""
        self.operation.set_running_state(False)
        if success:
            self.show_info(message)
        else:
            self.show_error(message)


class PreviewWidget(QWidget):
    """数据预览组件"""
    worksheet_changed = Signal(str)
    
    def __init__(self):
        super().__init__()
        self.workbook = None
        self.worksheet = None
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 工作表选择
        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(QLabel("工作表:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.currentTextChanged.connect(self.worksheet_changed.emit)
        sheet_layout.addWidget(self.sheet_combo)
        sheet_layout.addStretch()
        layout.addLayout(sheet_layout)
        
        # 数据预览表格
        self.table = QTableWidget()
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setMaximumHeight(200)  # 限制高度
        layout.addWidget(self.table)
        
    def update_workbook(self, workbook, worksheets):
        """更新工作簿"""
        self.workbook = workbook
        self.worksheets = worksheets
        
        # 更新工作表选择
        self.sheet_combo.clear()
        self.sheet_combo.addItems(worksheets)
        
        # 更新预览
        if worksheets:
            self.worksheet = workbook[worksheets[0]]
            self.update_worksheet(self.worksheet)
            
    def update_worksheet(self, worksheet):
        """更新工作表预览"""
        self.worksheet = worksheet
        
        # 限制显示行数和列数
        max_rows = min(20, worksheet.max_row)
        max_cols = min(10, worksheet.max_column)
        
        # 设置表格
        self.table.setRowCount(max_rows)
        self.table.setColumnCount(max_cols)
        
        # 设置列名
        col_labels = []
        for col in range(1, max_cols + 1):
            col_labels.append(openpyxl.utils.get_column_letter(col))
        self.table.setHorizontalHeaderLabels(col_labels)
        
        # 填充数据
        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                cell = worksheet.cell(row=row, column=col)
                value = str(cell.value) if cell.value is not None else ""
                self.table.setItem(row-1, col-1, QTableWidgetItem(value))


class FormulaApplicationWidget(QWidget):
    """公式应用配置组件"""
    
    def __init__(self):
        super().__init__()
        self.formula_configs = []
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 公式配置列表区域
        config_label = QLabel("公式配置列表:")
        layout.addWidget(config_label)
        
        self.config_scroll = QScrollArea()
        self.config_widget = QWidget()
        self.config_layout = QVBoxLayout(self.config_widget)
        self.config_scroll.setWidget(self.config_widget)
        self.config_scroll.setWidgetResizable(True)
        layout.addWidget(self.config_scroll)
        
        # 添加一个默认配置
        self.add_formula_config()
        
        # 添加配置按钮
        add_btn = QPushButton("添加公式配置")
        add_btn.clicked.connect(self.add_formula_config)
        layout.addWidget(add_btn)
        
    def add_formula_config(self):
        """添加一个公式配置"""
        config = FormulaConfigWidget(len(self.formula_configs))
        config.remove_requested.connect(self.remove_formula_config)
        self.formula_configs.append(config)
        self.config_layout.addWidget(config)
        
    def remove_formula_config(self, index):
        """移除公式配置"""
        if 0 <= index < len(self.formula_configs):
            config = self.formula_configs[index]
            self.config_layout.removeWidget(config)
            config.deleteLater()
            
            # 移除并更新索引
            self.formula_configs.pop(index)
            for i, conf in enumerate(self.formula_configs):
                conf.update_index(i)
                
    def get_formula_configs(self):
        """获取所有公式配置"""
        configs = []
        for config in self.formula_configs:
            config_data = config.get_config()
            if config_data:  # 只添加有效配置
                configs.append(config_data)
        return configs


class FormulaConfigWidget(QWidget):
    """单个公式配置组件"""
    remove_requested = Signal(int)
    
    def __init__(self, index):
        super().__init__()
        self.index = index
        self.setup_ui()
        
    def setup_ui(self):
        # 设置边框和标题样式
        self.setStyleSheet("QWidget { border: 1px solid #555; border-radius: 5px; margin: 5px; }")
        
        layout = QVBoxLayout(self)
        
        # 标题行
        title_layout = QHBoxLayout()
        title_label = QLabel(f"公式配置 #{self.index + 1}")
        title_label.setStyleSheet("font-weight: bold;")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        self.remove_btn = QPushButton("删除")
        self.remove_btn.setMaximumWidth(60)
        self.remove_btn.clicked.connect(lambda: self.remove_requested.emit(self.index))
        title_layout.addWidget(self.remove_btn)
        
        layout.addLayout(title_layout)
        
        # 公式类型选择
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("公式类型:"))
        self.formula_type_combo = QComboBox()
        self.formula_type_combo.addItems([
            "求和", "平均值", "计数", "最大值", "最小值", 
            "乘积", "条件计数", "文本连接", "自定义"
        ])
        self.formula_type_combo.currentTextChanged.connect(self.toggle_custom_formula)
        type_layout.addWidget(self.formula_type_combo)
        layout.addLayout(type_layout)
        
        # 自定义公式输入
        self.custom_formula_layout = QHBoxLayout()
        self.custom_formula_layout.addWidget(QLabel("自定义公式:"))
        self.custom_formula_input = QLineEdit()
        self.custom_formula_input.setPlaceholderText("例如: =SUM(A1:A10)*1.1")
        self.custom_formula_layout.addWidget(self.custom_formula_input)
        layout.addLayout(self.custom_formula_layout)
        
        # 默认隐藏自定义公式输入
        self.toggle_custom_formula(self.formula_type_combo.currentText())
        
        # 应用范围设置
        range_layout = QHBoxLayout()
        range_layout.addWidget(QLabel("数据范围:"))
        self.column_range_input = QLineEdit("A:A")
        self.column_range_input.setPlaceholderText("例如: A:A 或 A1:B10")
        range_layout.addWidget(self.column_range_input)
        layout.addLayout(range_layout)
        
        # 应用模式选择
        mode_layout = QHBoxLayout()
        mode_layout.addWidget(QLabel("应用模式:"))
        self.apply_mode_combo = QComboBox()
        self.apply_mode_combo.addItems(["单个单元格", "整列", "指定范围"])
        self.apply_mode_combo.currentTextChanged.connect(self.toggle_range_inputs)
        mode_layout.addWidget(self.apply_mode_combo)
        layout.addLayout(mode_layout)
        
        # 起始和结束行
        self.row_layout = QHBoxLayout()
        self.row_layout.addWidget(QLabel("起始行:"))
        self.start_row_input = QSpinBox()
        self.start_row_input.setMinimum(1)
        self.start_row_input.setValue(1)
        self.row_layout.addWidget(self.start_row_input)
        
        self.row_layout.addWidget(QLabel("结束行:"))
        self.end_row_input = QSpinBox()
        self.end_row_input.setMinimum(1)
        self.end_row_input.setValue(10)
        self.row_layout.addWidget(self.end_row_input)
        
        layout.addLayout(self.row_layout)
        
        # 默认隐藏行输入
        self.toggle_range_inputs(self.apply_mode_combo.currentText())
        
    def update_index(self, new_index):
        """更新索引"""
        self.index = new_index
        self.findChild(QLabel).setText(f"公式配置 #{self.index + 1}")
        
    def toggle_custom_formula(self, text):
        """根据公式类型切换自定义公式输入框可见性"""
        is_custom = text == "自定义"
        # 这里需要找到自定义公式输入布局，并设置其可见性
        for i in range(self.layout().count()):
            item = self.layout().itemAt(i)
            if item and item.layout() == self.custom_formula_layout:
                for j in range(item.layout().count()):
                    widget = item.layout().itemAt(j).widget()
                    if widget:
                        widget.setVisible(is_custom)
                break
                
    def toggle_range_inputs(self, text):
        """根据应用模式切换行输入框可见性"""
        is_range = text == "指定范围"
        # 这里需要找到行输入布局，并设置其可见性
        for i in range(self.layout().count()):
            item = self.layout().itemAt(i)
            if item and item.layout() == self.row_layout:
                for j in range(item.layout().count()):
                    widget = item.layout().itemAt(j).widget()
                    if widget:
                        widget.setVisible(is_range)
                break
                
    def get_config(self):
        """获取配置"""
        return {
            "formula_type": self.formula_type_combo.currentText(),
            "custom_formula": self.custom_formula_input.text(),
            "column_range": self.column_range_input.text(),
            "apply_mode": self.apply_mode_combo.currentText(),
            "start_row": self.start_row_input.value(),
            "end_row": self.end_row_input.value()
        }


if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication, QFileDialog, QScrollArea
    app = QApplication(sys.argv)
    window = ExcelFormulaWindow()
    window.show()
    sys.exit(app.exec())